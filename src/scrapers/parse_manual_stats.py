"""
PitchGuard — Manual Stats Excel Parser
========================================
Parses the manually filled Excel (Sheet 2 "Paste Data") where data is
arranged column-wise — one column per player.

Structure per player column:
    Row 1:  serial number
    Row 2:  player name
    Row 3:  TM ID
    Row 4:  "Paste Data here" header
    Then repeating blocks:
        season label   (e.g. "25/26")
        blank
        competition    (e.g. "Premier League")
        blank
        appearances    (e.g. "38")
        col2           (ignored)
        col3           (ignored)
        yellow cards   (ignored)
        2nd yellow     (ignored)
        red cards      (ignored)
        minutes        (e.g. "3,420'")
    ...repeats for each competition/season

Output: src/scrapers/data/raw/player_season_stats.csv
  Columns: player_tm_id, player_name, season, competition, appearances, minutes_played

Usage:
    python parse_manual_stats.py \
        --input  data/raw/player_stats_links.xlsx \
        --output src/scrapers/data/raw/player_season_stats.csv
"""

import argparse
import logging
import re
from pathlib import Path

import pandas as pd
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("parser")

# Seasons we care about
TARGET_SEASONS = {"19/20", "20/21", "21/22", "22/23", "23/24", "24/25", "25/26"}

# Regex patterns
SEASON_RE = re.compile(r"^\d{2}/\d{2}$")
MINUTES_RE = re.compile(r"^[\d,]+'$")
NUMERIC_RE = re.compile(r"^[\d,]+$")
DASH_RE = re.compile(r"^-+$")


def clean_val(val) -> str:
    """Convert cell value to stripped string, empty if None."""
    if val is None:
        return ""
    return str(val).strip()


def is_value(s: str) -> bool:
    return bool(MINUTES_RE.match(s) or NUMERIC_RE.match(s) or DASH_RE.match(s))


def parse_minutes(s: str) -> int | None:
    """Convert '3,420'' → 3420, '-' → None."""
    if not s or DASH_RE.match(s):
        return None
    cleaned = s.replace("'", "").replace(",", "").strip()
    try:
        return int(cleaned)
    except ValueError:
        return None


def parse_appearances(s: str) -> int | None:
    """Convert '38' → 38, '-' → None."""
    if not s or DASH_RE.match(s):
        return None
    cleaned = s.replace(",", "").strip()
    try:
        return int(cleaned)
    except ValueError:
        return None


def parse_player_column(cells: list[str], player_name: str, tm_id: str) -> list[dict]:
    """
    Parse one player's column of data.
    cells = list of all cell values in that column starting from the
    'Paste Data here' row onwards.
    """
    rows = []
    i = 0
    n = len(cells)

    # Target seasons filter
    target = TARGET_SEASONS

    while i < n:
        val = cells[i]

        # Look for a season label
        if SEASON_RE.match(val):
            season = val
            i += 1

            # Skip blank
            while i < n and cells[i] == "":
                i += 1

            # Next non-blank = competition name
            if i >= n:
                break
            competition = cells[i]
            i += 1

            # Skip blank
            while i < n and cells[i] == "":
                i += 1

            # Collect up to 7 value tokens
            values: list[str] = []
            while i < n and len(values) < 7:
                if is_value(cells[i]) or cells[i] == "":
                    values.append(cells[i])
                    i += 1
                else:
                    break

            # Only keep target seasons
            if season not in target:
                continue

            # Skip if competition looks like junk
            if not competition or competition.lower() in ("paste data here", ""):
                continue

            # appearances = first value token (index 0)
            appearances = None
            if values:
                appearances = parse_appearances(values[0])

            # minutes = last value token (index 6) — ends in '
            minutes = None
            for v in reversed(values):
                if MINUTES_RE.match(v):
                    minutes = parse_minutes(v)
                    break

            # Skip rows where both are None (no data)
            if appearances is None and minutes is None:
                continue

            rows.append(
                {
                    "player_tm_id": tm_id,
                    "player_name": player_name,
                    "season": season,
                    "competition": competition,
                    "appearances": appearances if appearances is not None else 0,
                    "minutes_played": minutes if minutes is not None else 0,
                }
            )
        else:
            i += 1

    return rows


def run(input_path: str, output_path: str) -> None:
    log.info(f"Loading Excel: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    # Find the "Paste Data" sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "paste" in name.lower() or "data" in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        # Fall back to second sheet
        if len(wb.sheetnames) >= 2:
            sheet_name = wb.sheetnames[1]
        else:
            sheet_name = wb.sheetnames[0]

    log.info(f"Using sheet: '{sheet_name}'")
    ws = wb[sheet_name]

    # Read all values into a 2D list
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([clean_val(cell) for cell in row])

    if not data:
        log.error("Sheet is empty.")
        return

    n_rows = len(data)
    n_cols = len(data[0]) if data else 0
    log.info(f"Sheet dimensions: {n_rows} rows × {n_cols} cols")

    # Find header rows
    # Row 0 = "No.", "Player Name", "TM ID", etc. headers
    # Row 1 = serial numbers
    # Row 2 = player names
    # Row 3 = TM IDs
    # Row 4 = "Paste Data here"
    # Row 5+ = actual data

    # Find which row has "Player Name"
    name_row_idx = None
    id_row_idx = None
    data_start_idx = None

    for r_idx, row in enumerate(data):
        row_lower = [c.lower() for c in row]
        if "player name" in row_lower:
            name_row_idx = r_idx
        if "tm id" in row_lower or "player tm id" in row_lower:
            id_row_idx = r_idx
        if "paste data here" in row_lower or any("paste" in c.lower() for c in row):
            data_start_idx = r_idx + 1
            break

    # Fallback: assume rows 1, 2, 3, start at 4
    if name_row_idx is None:
        name_row_idx = 1
    if id_row_idx is None:
        id_row_idx = 2
    if data_start_idx is None:
        data_start_idx = 4

    log.info(
        f"Name row: {name_row_idx}, ID row: {id_row_idx}, Data starts: {data_start_idx}"
    )

    # Find which columns have player data (non-empty name)
    player_cols = []
    for c_idx in range(1, n_cols):  # skip col 0 (label column)
        name = data[name_row_idx][c_idx] if c_idx < len(data[name_row_idx]) else ""
        tm_id = data[id_row_idx][c_idx] if c_idx < len(data[id_row_idx]) else ""
        if name and name.lower() not in ("player name", ""):
            player_cols.append((c_idx, name, str(tm_id).strip()))

    log.info(f"Found {len(player_cols)} player columns.")

    all_rows = []

    for c_idx, player_name, tm_id in player_cols:
        # Extract this player's column data from data_start_idx onward
        cells = []
        for r_idx in range(data_start_idx, n_rows):
            val = data[r_idx][c_idx] if c_idx < len(data[r_idx]) else ""
            cells.append(val)

        rows = parse_player_column(cells, player_name, tm_id)
        log.info(f"  {player_name} (id={tm_id}): {len(rows)} season-competition rows")
        all_rows.extend(rows)

    if not all_rows:
        log.warning("No rows parsed. Check sheet structure.")
        return

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    # If existing file, merge (don't overwrite players not in this Excel)
    if out.exists() and out.stat().st_size > 0:
        existing = pd.read_csv(out, dtype=str)
        new_df = pd.DataFrame(all_rows)
        # Remove existing rows for players we just re-parsed
        parsed_ids = {r["player_tm_id"] for r in all_rows}
        existing = existing[~existing["player_tm_id"].astype(str).isin(parsed_ids)]
        combined = pd.concat([existing, new_df], ignore_index=True)
        combined.to_csv(out, index=False, encoding="utf-8-sig")
        log.info("\n✅ Merged with existing file.")
        log.info(f"   Total rows: {len(combined)}")
    else:
        df = pd.DataFrame(all_rows)
        df.to_csv(out, index=False, encoding="utf-8-sig")
        log.info("\n✅ Saved fresh.")
        log.info(f"   Total rows: {len(df)}")

    log.info(f"   Output: {out}")
    log.info(f"   Players parsed: {len(player_cols)}")
    log.info(f"   Rows parsed: {len(all_rows)}")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--input", "-i", default="data/raw/player_stats_links.xlsx")
    p.add_argument(
        "--output", "-o", default="src/scrapers/data/raw/player_season_stats.csv"
    )
    args = p.parse_args()
    run(args.input, args.output)
